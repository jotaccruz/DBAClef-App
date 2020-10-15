from openpyxl import Workbook, cell
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, NamedStyle,Font
from openpyxl.styles import Color, Border, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
#from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

def as_text(value):
    if value is None:
        return ""
    return str(value)

def xlsxGen(excelf,section,tree,tree_name,tree_dic):
    #Style
    red_fill = PatternFill(bgColor="FFC7CE")
    orange_fill = PatternFill(bgColor="F5E45E")
    dxfcritical = DifferentialStyle(fill=red_fill)
    dxfwarning = DifferentialStyle(fill=orange_fill)

    rcritical = Rule(type="expression", dxf=dxfcritical, stopIfTrue=True)
    rcritical1 = Rule(type="expression", dxf=dxfcritical, stopIfTrue=True)
    rcritical2 = Rule(type="expression", dxf=dxfcritical, stopIfTrue=True)
    rcritical3 = Rule(type="expression", dxf=dxfcritical, stopIfTrue=True)
    rcritical4 = Rule(type="expression", dxf=dxfcritical, stopIfTrue=True)
    rcritical5 = Rule(type="expression", dxf=dxfcritical, stopIfTrue=True)

    rwarning = Rule(type="expression", dxf=dxfwarning, stopIfTrue=True)
    rwarning1 = Rule(type="expression", dxf=dxfwarning, stopIfTrue=True)
    rwarning2 = Rule(type="expression", dxf=dxfwarning, stopIfTrue=True)
    rwarning3 = Rule(type="expression", dxf=dxfwarning, stopIfTrue=False)
    rwarning4 = Rule(type="expression", dxf=dxfwarning, stopIfTrue=False)
    rwarning5 = Rule(type="expression", dxf=dxfwarning, stopIfTrue=False)

    rcritical.formula = ['$C1="Disabled"']
    rcritical1.formula = ['$E1="Take Care"']
    rcritical2.formula = ['$B1="Missing"']
    rcritical3.formula = ['$C1="Missing"']
    rcritical4.formula = ['$C1="BUILTIN\\Users"']
    rcritical5.formula = ['$C1="NT AUTHORITY\\SYSTEM"']

    rwarning.formula = ['$F1="Disabled"']
    #rwarning1.formula = ['$F1<>"sa"']
    rwarning2.formula = ['$D1="Required"']
    rwarning3.formula = ['$E1="OFF"']
    rwarning4.formula = ['$AD1="N"']
    rwarning5.formula = ['$AE1="N"']


    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")

    #-----

    wb=excelf
    ws = wb.active
    if ("Sheet" in wb.sheetnames):
        ws.title = section
    else:
        if section in wb.sheetnames:
            ws = wb[section]
        else:
            ws = wb.create_sheet(title=section)

    columnslist=tree['columns']
    row=1
    col=1
    for titles in columnslist:
        ws.cell (row=row,column=col,value=titles)
        ws.cell (row,col).border = Border(top=thin, left=thin, right=thin, \
        bottom=thin)
        col=col+1

    row=3
    for lists in tree.get_children():
        col=1
        for values in tree.item(lists)['values']:
            ws.cell (column=col,row=row,value=values)
            ws.conditional_formatting.add("A1:Z500", rcritical)
            ws.conditional_formatting.add("A1:Z500", rcritical1)
            ws.conditional_formatting.add("A1:Z500", rcritical2)
            ws.conditional_formatting.add("A1:Z500", rcritical3)
            ws.conditional_formatting.add("A1:Z500", rcritical4)
            ws.conditional_formatting.add("A1:Z500", rcritical5)

            ws.conditional_formatting.add("A1:Z500", rwarning)
            #ws.conditional_formatting.add("A1:Z500", rwarning1)
            ws.conditional_formatting.add("A1:Z500", rwarning2)
            ws.conditional_formatting.add("A1:Z500", rwarning3)
            ws.conditional_formatting.add("A1:AG500", rwarning4)
            ws.conditional_formatting.add("A1:AG500", rwarning5)

            col=col+1
        row=row+1

    for key in tree_dic:
        #print (key)
        if key==tree_name:
            i=1
            for values in tree_dic[key]:
                #print (get_column_letter(i))
                ws.column_dimensions[get_column_letter(i)].width = values
                i=i+1

        #print(titles)

    #ws['A1']="Hola"
    #wb.save('test.xlsx')

    #Applying Style
    return wb
